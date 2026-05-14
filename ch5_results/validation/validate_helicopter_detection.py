#!/usr/bin/env python3
"""
Validate helicopter event detection against Master Flight Report.

Compares detected events in stage3_helicopter_events against actual offshore
flight logs from the industry partner. Computes recall (what fraction of real
offshore flights we detected) and analyses false-positive rate.

Usage:
    python scripts/validate_helicopter_detection.py [--score-min 25]
"""

import argparse
import datetime as dt
from collections import defaultdict

import openpyxl
import psycopg2
from zoneinfo import ZoneInfo

# Config

FLIGHT_REPORT_PRIMARY  = "/mnt/e/data_lake/helicopters/validation/Master Flight Report.xlsx"
FLIGHT_REPORT_FALLBACK = "/mnt/d/thesis/Master Flight Report.xlsx"
SHEETS_2024    = ["MVY 2024", "OQU 2024"]
SHEETS_2025_H1 = ["MVY 2025", "OQU 2025"]   # filtered to dof < 2025-07-01

# v2 is_offshore rule: MFR rows with DEP=DES whose duration exceeds this threshold
# are treated as offshore round-trips.  Justification: empirical KOQU↔wind-farm
# round-trip durations cluster in 30-60 min (single-direction transit 17-30 min;
# see parameter_justification/02_same_airport_duration.png and 03_transit_time_table.tsv).
SAME_AIRPORT_OFFSHORE_MIN_MINUTES = 30

# Canonical → sheet header label.  Sheets differ in column order / presence:
#   MVY 2024, OQU 2024, MVY 2025: 9 cols, no "Mission"
#   OQU 2025:                     9 cols, "Mission" inserted at col 2,
#                                 "BLK On" dropped from col 9
HEADER_LABELS = {
    "reg":     "A/C REG",
    "mission": "Mission",
    "dep":     "DEP",
    "des":     "DES",
    "ldg_num": "LDG #",
    "dof":     "DOF",
    "blk_off": "BLK Off",
    "to":      "T/O",
    "ldg":     "LDG",
    "blk_on":  "BLK On",
}


def _resolve_flight_report() -> str:
    """Prefer the primary path; fall back to the local copy when the data
    lake drive is offline. WSL drvfs raises OSError (not False) on a
    disconnected mount, so guard with try/except."""
    from pathlib import Path
    try:
        if Path(FLIGHT_REPORT_PRIMARY).exists():
            return FLIGHT_REPORT_PRIMARY
    except OSError:
        pass
    return FLIGHT_REPORT_FALLBACK


FLIGHT_REPORT = _resolve_flight_report()

DB = dict(host="localhost", port=5432, dbname="windfarm",
          user="thesis", password="thesis2026")

# Known airport codes (anything not in this set is offshore)
AIRPORT_CODES = {
    "KOQU", "KMVY", "KACK", "KACY", "KEWB", "KGON", "KHYA", "KLDG", "KPVD",
    "MVY", "koqu", "6N5", "BKL1",
}

# Tail number → ICAO24 mapping (verified via OpenSky aircraft metadata API
# at /api/metadata/aircraft/icao/{hex}). The earlier heuristic mapping (from
# airport frequency correlation) was rotated by one and has been corrected.
TAIL_TO_ICAO24 = {
    "N691HS": "a92f2d",
    "N692HS": "a932e4",
    "N693HS": "a9369b",
    "N694HS": "a93a52",  # newly identified; AW169 added to HSU fleet during 2025
}
# Reverse mapping
ICAO24_TO_TAIL = {v: k for k, v in TAIL_TO_ICAO24.items()}

# ICAO24 hex codes for the known helicopter fleet — derived from TAIL_TO_ICAO24.
# (The previously-included a87968 was N645SK, a Raytheon Beechcraft owned by
# Knafel Aviation, not an HSU helicopter; it has been removed.)
KNOWN_FLEET = set(TAIL_TO_ICAO24.values())

ET = ZoneInfo("America/New_York")  # handles EDT/EST automatically
UTC = ZoneInfo("UTC")


# Parse flight report

def parse_time(val) -> dt.time | None:
    """Parse time from various Excel formats."""
    if val is None:
        return None
    if isinstance(val, dt.time):
        return val
    if isinstance(val, dt.timedelta):
        total_sec = int(val.total_seconds())
        h, m = divmod(total_sec // 60, 60)
        return dt.time(h % 24, m)
    if isinstance(val, str):
        val = val.strip()
        if ":" in val:
            parts = val.split(":")
            return dt.time(int(parts[0]), int(parts[1]))
    return None


def _build_header_map(ws) -> dict[str, int]:
    """Read the header row (row 2) and return {canonical: column_index}."""
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    name_to_idx: dict[str, int] = {}
    for canonical, label in HEADER_LABELS.items():
        for i, cell in enumerate(header_row):
            if cell == label:
                name_to_idx[canonical] = i
                break
    return name_to_idx


def load_flights(sheets: list[str],
                 date_min: dt.date | None = None,
                 date_max: dt.date | None = None) -> list[dict]:
    """Load flights from the named sheets, optionally filtered to a date range.

    Header-driven: each sheet is parsed by looking up canonical fields
    (DOF, DEP, DES, T/O, LDG, BLK Off/On, A/C REG, Mission) in the header
    row at row 2.  This handles the OQU 2025 schema variant which inserts
    a Mission column and drops BLK On.
    """
    wb = openpyxl.load_workbook(FLIGHT_REPORT, read_only=True, data_only=True)
    flights = []

    for sheet_name in sheets:
        ws = wb[sheet_name]
        cols = _build_header_map(ws)
        if "dof" not in cols or "dep" not in cols or "des" not in cols:
            print(f"  WARN: {sheet_name} missing canonical fields, skipping")
            continue

        for row in ws.iter_rows(min_row=3, values_only=True):
            def get(field):
                idx = cols.get(field)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            dof = get("dof")
            dep = get("dep")
            des = get("des")
            if dof is None or dep is None or des is None:
                continue

            # Normalize
            dep = str(dep).strip()
            des = str(des).strip()
            reg = str(get("reg")).strip() if get("reg") else None
            mission = str(get("mission")).strip() if get("mission") else None

            # Parse date
            if isinstance(dof, dt.datetime):
                flight_date = dof.date()
            else:
                continue

            # Date filter
            if date_min and flight_date < date_min:
                continue
            if date_max and flight_date > date_max:
                continue

            # Parse takeoff and landing times (local) — needed for the v2 same-airport
            # offshore-duration rule below.
            to_local = parse_time(get("to"))
            ldg_local = parse_time(get("ldg"))

            # Convert to UTC
            to_utc = None
            ldg_utc = None
            if to_local:
                local_dt = dt.datetime.combine(flight_date, to_local, tzinfo=ET)
                to_utc = local_dt.astimezone(UTC)
            if ldg_local:
                local_dt = dt.datetime.combine(flight_date, ldg_local, tzinfo=ET)
                ldg_utc = local_dt.astimezone(UTC)

            # Is this an offshore flight?
            dep_upper = dep.upper()
            des_upper = des.upper()
            dep_is_airport = dep_upper in {a.upper() for a in AIRPORT_CODES}
            des_is_airport = des_upper in {a.upper() for a in AIRPORT_CODES}
            is_offshore = not dep_is_airport or not des_is_airport

            # v2 rule: a same-airport (DEP=DES) row whose duration exceeds
            # SAME_AIRPORT_OFFSHORE_MIN_MINUTES is treated as an offshore round-trip.
            # The MFR encodes some offshore sorties as "KOQU → KOQU" (departs and
            # returns to the same airport with no separate offshore-destination row);
            # empirically (parameter_justification/02_same_airport_duration.png)
            # ~75% of same-airport legs fall in 30-60 min, exactly the KOQU↔wind-farm
            # round-trip envelope.
            if not is_offshore and dep == des and to_utc and ldg_utc:
                dur_min = (ldg_utc - to_utc).total_seconds() / 60
                if dur_min > SAME_AIRPORT_OFFSHORE_MIN_MINUTES:
                    is_offshore = True

            # OQU 2025 has an explicit Mission column; treat "Offshore" / "Hoist" /
            # vessel-named missions as offshore even if both endpoints are airports
            # and the duration rule didn't fire.
            if mission and not is_offshore:
                m = mission.lower()
                if "offshore" in m or "hoist" in m or "windfarm" in m:
                    is_offshore = True

            flights.append({
                "reg": reg,
                "dep": dep,
                "des": des,
                "mission": mission,
                "date": flight_date,
                "to_utc": to_utc,
                "ldg_utc": ldg_utc,
                "is_offshore": is_offshore,
                "sheet": sheet_name,
            })

    wb.close()
    return flights


def load_flights_2024() -> list[dict]:
    """Load 2024 flights only (backward-compat wrapper)."""
    return load_flights(SHEETS_2024,
                        date_min=dt.date(2024, 1, 1),
                        date_max=dt.date(2024, 12, 31))


def load_flights_2025_h1() -> list[dict]:
    """Load 2025 H1 flights (Jan 1 – Jun 30) from the 2025 sheets."""
    return load_flights(SHEETS_2025_H1,
                        date_min=dt.date(2025, 1, 1),
                        date_max=dt.date(2025, 6, 30))


def load_flights_combined() -> list[dict]:
    """Load 2024 + 2025 H1 combined."""
    return load_flights_2024() + load_flights_2025_h1()


# Load detected events

def load_detected_events(score_min: float,
                         date_min: dt.date = dt.date(2024, 1, 1),
                         date_max: dt.date = dt.date(2025, 1, 1)) -> list[dict]:
    """Load helicopter events from database (turbine visits + SOV hoists).

    Date range is half-open [date_min, date_max) and must align with the MFR
    corpus loaded by load_flights().
    """
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Turbine visit events — date-scoped to match the MFR corpus.
    cur.execute("""
        SELECT icao24, visit_start, visit_end, turbine_code, project_name,
               duration_minutes, score, n_positions
        FROM stage3_helicopter_events
        WHERE score >= %s
          AND visit_start >= %s
          AND visit_start <  %s
        ORDER BY visit_start
    """, (score_min, date_min, date_max))

    events = []
    for row in cur.fetchall():
        events.append({
            "icao24": row[0],
            "visit_start": row[1],
            "visit_end": row[2],
            "turbine_code": row[3],
            "project_name": row[4],
            "duration_min": row[5],
            "score": row[6],
            "n_positions": row[7],
            "source": "turbine",
        })

    # SOV hoist events (from stage4_sov_interactions) — same date filter
    cur.execute("""
        SELECT asset_id, interaction_start, interaction_end,
               sov_name, project_name, duration_minutes, score
        FROM stage4_sov_interactions
        WHERE asset_type = 'helicopter'
          AND interaction_type = 'hoist'
          AND score >= %s
          AND interaction_start >= %s
          AND interaction_start <  %s
        ORDER BY interaction_start
    """, (score_min, date_min, date_max))

    for row in cur.fetchall():
        events.append({
            "icao24": row[0],
            "visit_start": row[1],
            "visit_end": row[2],
            "turbine_code": f"SOV:{row[3]}",  # mark as SOV interaction
            "project_name": row[4],
            "duration_min": row[5],
            "score": row[6],
            "n_positions": None,
            "source": "sov_hoist",
        })

    conn.close()
    events.sort(key=lambda e: e["visit_start"])
    return events


# Matching logic

def match_flights_to_events(flights: list[dict], events: list[dict],
                            time_tolerance_min: int = 120):
    """
    Match actual offshore flights to detected events.

    Strategy: For each actual offshore flight, find detected events on the same
    UTC date (±1 day for timezone edge cases) whose visit window overlaps with
    the flight's takeoff-to-landing window (expanded by tolerance).

    The 120-minute tolerance accounts for:
    - Multi-turbine sorties where last event may be 60-90 min after takeoff
    - ADS-B timestamp drift vs flight log local-time conversion
    - Hoist operations en route (to/from vessels) extending total flight duration

    Aircraft identity filtering: only match events from the same helicopter
    (using TAIL_TO_ICAO24 mapping). Events from unmapped aircraft match any flight.

    Returns matched flights, unmatched flights, and unmatched events.
    """
    # Index events by UTC date
    events_by_date = defaultdict(list)
    for e in events:
        d = e["visit_start"].date()
        events_by_date[d].append(e)
        # Also index ±1 day for edge cases
        events_by_date[d - dt.timedelta(days=1)].append(e)
        events_by_date[d + dt.timedelta(days=1)].append(e)

    tolerance = dt.timedelta(minutes=time_tolerance_min)
    matched_flights = []
    unmatched_flights = []
    matched_event_ids = set()

    for f in flights:
        if not f["is_offshore"]:
            continue

        candidates = events_by_date.get(f["date"], [])
        if not candidates:
            unmatched_flights.append(f)
            continue

        # Time-based matching (aircraft identity not filtered — ICAO24↔tail
        # mapping is provisional and unconfirmed; filtering would reduce recall
        # without improving precision since all aircraft serve the same operator)
        best_matches = []
        for e in candidates:
            # Only match events from the known fleet
            if e["icao24"] not in KNOWN_FLEET:
                continue

            # If we have takeoff/landing times, use time overlap
            if f["to_utc"] and f["ldg_utc"]:
                flight_start = f["to_utc"] - tolerance
                flight_end = f["ldg_utc"] + tolerance
                if (e["visit_start"] <= flight_end and
                        e["visit_end"] >= flight_start):
                    best_matches.append(e)
            else:
                # Date-only match
                if e["visit_start"].date() == f["date"]:
                    best_matches.append(e)

        if best_matches:
            matched_flights.append((f, best_matches))
            for e in best_matches:
                eid = (e["icao24"], e["visit_start"], e["turbine_code"])
                matched_event_ids.add(eid)
        else:
            unmatched_flights.append(f)

    # Find unmatched events
    unmatched_events = []
    for e in events:
        eid = (e["icao24"], e["visit_start"], e["turbine_code"])
        if eid not in matched_event_ids:
            unmatched_events.append(e)

    return matched_flights, unmatched_flights, unmatched_events


# Canonical (strict) matching: inter-leg dwell windows
#
# This is the methodologically tightest matching mode the data supports.
# Compared to the per-leg scheme above:
#   - The matching unit is the *offshore presence interval* (when the helicopter
#     is known to be at an offshore location), not the airport-to-airport leg
#     window.  Three window kinds:
#       (a) inter-leg: [LDG of leg N, T/O of leg N+1] when leg_N.DES is
#           non-airport.  The MFR directly tells us the helicopter was at
#           leg_N.DES during this interval.
#       (b) same-airport: [T/O, LDG] for KOQU→KOQU rows whose duration exceeds
#           SAME_AIRPORT_OFFSHORE_MIN_MINUTES.  Cannot be decomposed into
#           transit/dwell/transit without per-destination info, so we keep the
#           full envelope (transit included).
#       (c) last-leg-arrival: [LDG, end-of-day] when the last leg ends at an
#           offshore destination.
#   - Aircraft identity is enforced strictly via TAIL_TO_ICAO24 (verified via
#     OpenSky aircraft-metadata API), not just KNOWN_FLEET membership.
#   - Match condition is pure interval overlap (no tolerance buffer).
#
# Recall is reported per (date, tail) pair (the natural operational unit for
# daily fleet activity); precision over events at score >= score_min.

def build_dwell_windows(flights: list[dict]) -> list[dict]:
    """Construct offshore-presence dwell windows from MFR flight rows.

    Returns a list of dicts: {date, tail, window_start, window_end, kind, location, contributing_legs}
    """
    from collections import defaultdict
    AIRPORTS_U = {a.upper() for a in AIRPORT_CODES}
    by_td: dict = defaultdict(list)
    for f in flights:
        if f.get("to_utc") and f.get("ldg_utc") and f.get("reg"):
            by_td[(f["date"], f["reg"])].append(f)

    windows: list[dict] = []
    for (d, tail), legs in by_td.items():
        legs.sort(key=lambda f: f["to_utc"])
        # (a) inter-leg dwell
        for i in range(len(legs) - 1):
            cur, nxt = legs[i], legs[i + 1]
            if cur["des"].upper() in AIRPORTS_U:
                continue
            if cur["ldg_utc"] < nxt["to_utc"]:
                windows.append({
                    "date": d, "tail": tail,
                    "window_start": cur["ldg_utc"], "window_end": nxt["to_utc"],
                    "kind": "inter-leg", "location": cur["des"],
                    "contributing_legs": [cur, nxt],
                })
        # (b) same-airport long round-trip
        for f in legs:
            if (f["dep"] == f["des"] and f["to_utc"] and f["ldg_utc"]
                    and (f["ldg_utc"] - f["to_utc"]).total_seconds() / 60
                        > SAME_AIRPORT_OFFSHORE_MIN_MINUTES):
                windows.append({
                    "date": d, "tail": tail,
                    "window_start": f["to_utc"], "window_end": f["ldg_utc"],
                    "kind": "same-airport", "location": f["dep"] + " (round-trip)",
                    "contributing_legs": [f],
                })
        # (c) last leg arriving offshore
        if legs:
            last = legs[-1]
            if last["des"].upper() not in AIRPORTS_U:
                eod = dt.datetime(d.year, d.month, d.day, 23, 59, 59,
                                  tzinfo=last["ldg_utc"].tzinfo)
                if last["ldg_utc"] < eod:
                    windows.append({
                        "date": d, "tail": tail,
                        "window_start": last["ldg_utc"], "window_end": eod,
                        "kind": "last-leg-arrival", "location": last["des"],
                        "contributing_legs": [last],
                    })
    return windows


def match_flights_strict_dwell(flights: list[dict], events: list[dict]):
    """Canonical (strict + dwell + zero-tolerance) matching.

    Returns:
      matched_pairs       : list of (window, event) tuples
      matched_pair_keys   : set of (date, tail) pairs with at least one match
      matched_event_keys  : set of (icao24, visit_start, turbine_code) keys
      unmatched_pairs     : set of offshore (date, tail) pairs with no match
      unmatched_events    : list of KNOWN_FLEET events with no window match
    """
    from collections import defaultdict
    windows = build_dwell_windows(flights)
    offshore_pairs = {(f["date"], f["reg"]) for f in flights if f.get("is_offshore")}

    events_by_date_icao: dict = defaultdict(list)
    for e in events:
        if e["icao24"] not in KNOWN_FLEET:
            continue
        d = e["visit_start"].date()
        for off in (-1, 0, 1):
            events_by_date_icao[(d + dt.timedelta(days=off), e["icao24"])].append(e)

    def ek(e):
        return (e["icao24"], e["visit_start"], e["turbine_code"])

    matched_pairs: list = []
    matched_pair_keys: set = set()
    matched_event_keys: set = set()
    for w in windows:
        expected = TAIL_TO_ICAO24.get(w["tail"])
        if not expected:
            continue
        for e in events_by_date_icao.get((w["date"], expected), []):
            # pure interval overlap, no tolerance
            if e["visit_start"] <= w["window_end"] and e["visit_end"] >= w["window_start"]:
                matched_pairs.append((w, e))
                matched_pair_keys.add((w["date"], w["tail"]))
                matched_event_keys.add(ek(e))

    unmatched_pairs = offshore_pairs - matched_pair_keys
    unmatched_events = [e for e in events
                        if ek(e) not in matched_event_keys
                           and e["icao24"] in KNOWN_FLEET]
    return matched_pairs, matched_pair_keys, matched_event_keys, unmatched_pairs, unmatched_events


# Sortie-level grouping and matching
#
# A sortie is one out-and-back trip by a single aircraft on a single date.
# The MFR records each *leg* (one takeoff -> one landing) as its own row,
# and `ldg_num` is unreliable in this dataset (every row reads as 1).  We
# group legs into sorties by walking each (aircraft, date) bucket in
# takeoff-time order and closing a sortie whenever a leg lands at a base
# airport.  The sortie envelope is min(takeoff) -> max(landing) across
# its legs, which is the natural matching unit because the helicopter is
# at the turbine *between* legs --- exactly inside that envelope.
#
# Sortie matching uses a small tolerance (default 30 min for clock drift)
# rather than the 60-120 min the per-leg scheme needs to absorb in-between
# dwell time.

def group_legs_into_sorties(legs: list[dict]) -> list[dict]:
    """Walk each (reg, date) bucket and split into sorties at each
    return-to-base. Returns one dict per sortie."""
    BASE_CODES_UPPER = {a.upper() for a in AIRPORT_CODES}
    by_aircraft_date: dict = defaultdict(list)
    for f in legs:
        if f.get("date") is None:
            continue
        by_aircraft_date[(f.get("reg"), f["date"])].append(f)

    sorties: list[dict] = []
    for (reg, date), day_legs in by_aircraft_date.items():
        # Skip aircraft we can't sort (no takeoff time)
        day_legs = [l for l in day_legs if l.get("to_utc") is not None]
        day_legs.sort(key=lambda l: l["to_utc"])

        current: list[dict] = []
        for leg in day_legs:
            current.append(leg)
            des_upper = str(leg.get("des", "")).strip().upper()
            if des_upper in BASE_CODES_UPPER:
                sorties.append(_make_sortie(current))
                current = []
        if current:  # trailing legs that never returned to a base
            sorties.append(_make_sortie(current))
    return sorties


def _make_sortie(legs: list[dict]) -> dict:
    takeoffs = [l["to_utc"] for l in legs if l.get("to_utc")]
    landings = [l["ldg_utc"] for l in legs if l.get("ldg_utc")]
    return {
        "reg":               legs[0].get("reg"),
        "date":              legs[0]["date"],
        "first_takeoff_utc": min(takeoffs) if takeoffs else None,
        "last_landing_utc":  max(landings) if landings else None,
        "n_legs":            len(legs),
        "is_offshore":       any(l["is_offshore"] for l in legs),
        "departure":         legs[0].get("dep"),
        "destinations":      [l.get("des") for l in legs],
        "legs":              legs,
    }


def match_sorties_to_events(sorties: list[dict], events: list[dict],
                            time_tolerance_min: int = 30):
    """Match detected events to sorties by envelope overlap.

    Returns (matched_sorties, unmatched_sorties, unmatched_events) where
    matched_sorties is a list of (sortie, [event, ...]) tuples.
    """
    events_by_date: dict = defaultdict(list)
    for e in events:
        d = e["visit_start"].date()
        events_by_date[d].append(e)
        events_by_date[d - dt.timedelta(days=1)].append(e)
        events_by_date[d + dt.timedelta(days=1)].append(e)

    tolerance = dt.timedelta(minutes=time_tolerance_min)
    matched_sorties = []
    unmatched_sorties = []
    matched_event_ids = set()

    for s in sorties:
        if not s["is_offshore"]:
            continue
        if s["first_takeoff_utc"] is None or s["last_landing_utc"] is None:
            unmatched_sorties.append(s)
            continue

        candidates = events_by_date.get(s["date"], [])
        envelope_start = s["first_takeoff_utc"] - tolerance
        envelope_end   = s["last_landing_utc"]  + tolerance

        matches = []
        for e in candidates:
            if e["icao24"] not in KNOWN_FLEET:
                continue
            # Event window overlap with sortie envelope
            if e["visit_start"] <= envelope_end and e["visit_end"] >= envelope_start:
                matches.append(e)

        if matches:
            matched_sorties.append((s, matches))
            for e in matches:
                matched_event_ids.add(
                    (e["icao24"], e["visit_start"], e["turbine_code"]))
        else:
            unmatched_sorties.append(s)

    unmatched_events = [
        e for e in events
        if (e["icao24"], e["visit_start"], e["turbine_code"]) not in matched_event_ids
    ]
    return matched_sorties, unmatched_sorties, unmatched_events


def print_sortie_report(sorties: list[dict], events: list[dict],
                         matched, unmatched_sorties, unmatched_events,
                         score_min: float):
    offshore_sorties = [s for s in sorties if s["is_offshore"]]
    n_off = len(offshore_sorties)
    n_matched = len(matched)
    n_unmatched_events = len(unmatched_events)
    n_matched_events = len(events) - n_unmatched_events

    print(f"\n{'SORTIE-LEVEL VALIDATION':-^70}")
    print(f"  Total sorties (after grouping legs):  {len(sorties):>6}")
    print(f"  Offshore sorties:                     {n_off:>6}")
    print(f"  Single-leg sorties (no return):       "
          f"{sum(1 for s in sorties if s['n_legs']==1):>6}")
    print(f"  Multi-leg sorties:                    "
          f"{sum(1 for s in sorties if s['n_legs']>1):>6}")
    print(f"  Median legs per offshore sortie:      "
          f"{(sorted(s['n_legs'] for s in offshore_sorties) or [0])[n_off//2]:>6}")

    recall    = 100 * n_matched / n_off if n_off else 0
    precision = 100 * n_matched_events / len(events) if events else 0
    f1 = 2 * recall * precision / (recall + precision) if (recall + precision) > 0 else 0
    print(f"\n  Matched offshore sorties:             {n_matched:>5} / {n_off}")
    print(f"  Sortie-level recall:                  {recall:>5.1f}%")
    print(f"  Detected events matched:              {n_matched_events:>5} / {len(events)}")
    print(f"  Event-level precision:                {precision:>5.1f}%")
    print(f"  F1-score:                             {f1:>5.1f}%")


# Reporting

def print_report(flights, events, matched, unmatched_flights, unmatched_events,
                 score_min):
    offshore = [f for f in flights if f["is_offshore"]]
    ferry = [f for f in flights if not f["is_offshore"]]

    # Unique flight-days for offshore flights
    offshore_days = set()
    for f in offshore:
        offshore_days.add(f["date"])

    # Unique flight-days detected
    detected_days = set()
    for e in events:
        detected_days.add(e["visit_start"].date())

    matched_days = set()
    for f, _ in matched:
        matched_days.add(f["date"])

    print("=" * 70)
    print(f"  HELICOPTER DETECTION VALIDATION (score >= {score_min})")
    print("=" * 70)

    print(f"\n{'FLIGHT LOG SUMMARY':-^70}")
    print(f"  Total 2024 flights:     {len(flights):>6}")
    print(f"  Offshore flights:       {len(offshore):>6}")
    print(f"  Ferry flights:          {len(ferry):>6}")
    print(f"  Unique offshore dates:  {len(offshore_days):>6}")

    print(f"\n{'DETECTION SUMMARY':-^70}")
    print(f"  Detected events:        {len(events):>6}")
    print(f"  Unique detected dates:  {len(detected_days):>6}")

    # Source breakdown
    turbine_events = [e for e in events if e.get("source") == "turbine"]
    sov_events     = [e for e in events if e.get("source") == "sov_hoist"]
    print(f"  Turbine visit events:   {len(turbine_events):>6}")
    print(f"  SOV hoist events:       {len(sov_events):>6}")

    # Fleet breakdown
    fleet_events = [e for e in events if e["icao24"] in KNOWN_FLEET]
    other_events = [e for e in events if e["icao24"] not in KNOWN_FLEET]
    print(f"  From known fleet:       {len(fleet_events):>6} ({len(fleet_events)/len(events)*100:.0f}%)")
    print(f"  From other aircraft:    {len(other_events):>6}")

    print(f"\n{'MATCHING RESULTS':-^70}")
    print(f"  Matched offshore flights:   {len(matched):>5} / {len(offshore)}")
    print(f"  Unmatched offshore flights: {len(unmatched_flights):>5}")
    recall = len(matched) / len(offshore) * 100 if offshore else 0
    print(f"  Flight-level recall:        {recall:>5.1f}%")

    day_recall = len(matched_days) / len(offshore_days) * 100 if offshore_days else 0
    print(f"\n  Matched offshore dates:     {len(matched_days):>5} / {len(offshore_days)}")
    print(f"  Day-level recall:           {day_recall:>5.1f}%")

    # Event-level precision: events matched to a flight / total detected events
    n_matched_events = len(events) - len(unmatched_events)
    precision = n_matched_events / len(events) * 100 if events else 0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) > 0 else 0

    print(f"\n  {'PRECISION / RECALL / F1':-^66}")
    print(f"  Events matched to a flight: {n_matched_events:>5} / {len(events)}")
    print(f"  Event-level precision:      {precision:>5.1f}%")
    print(f"  Flight-level recall:        {recall:>5.1f}%")
    print(f"  F1-score:                   {f1:>5.1f}%")
    print(f"  Unmatched events:           {len(unmatched_events):>5}")

    # Events per matched flight (expect ~1.6 for multi-turbine visits)
    total_matched_events = sum(len(evts) for _, evts in matched)
    events_per_flight = total_matched_events / len(matched) if matched else 0
    print(f"\n  Events per matched flight:  {events_per_flight:>5.1f} "
          f"(multi-turbine visits per sortie)")

    # Monthly breakdown
    print(f"\n{'MONTHLY RECALL':-^70}")
    monthly_offshore = defaultdict(int)
    monthly_matched = defaultdict(int)
    for f in offshore:
        monthly_offshore[f["date"].month] += 1
    for f, _ in matched:
        monthly_matched[f["date"].month] += 1

    print(f"  {'Month':>7}  {'Offshore':>8}  {'Matched':>7}  {'Recall':>6}")
    for m in sorted(set(list(monthly_offshore.keys()) + list(monthly_matched.keys()))):
        off = monthly_offshore[m]
        mat = monthly_matched[m]
        r = mat / off * 100 if off > 0 else 0
        print(f"  {m:>7}  {off:>8}  {mat:>7}  {r:>5.1f}%")

    # Score distribution of matched vs unmatched events
    print(f"\n{'SCORE DISTRIBUTION':-^70}")
    matched_scores = []
    for _, evts in matched:
        for e in evts:
            matched_scores.append(e["score"])
    unmatched_scores = [e["score"] for e in unmatched_events]

    if matched_scores:
        matched_scores.sort()
        n = len(matched_scores)
        print(f"  Matched events:   median={matched_scores[n//2]:.1f}  "
              f"P25={matched_scores[n//4]:.1f}  P75={matched_scores[3*n//4]:.1f}")
    if unmatched_scores:
        unmatched_scores.sort()
        n = len(unmatched_scores)
        print(f"  Unmatched events: median={unmatched_scores[n//2]:.1f}  "
              f"P25={unmatched_scores[n//4]:.1f}  P75={unmatched_scores[3*n//4]:.1f}")

    # Unmatched flight details (sample)
    if unmatched_flights:
        print(f"\n{'SAMPLE UNMATCHED FLIGHTS (first 15)':-^70}")
        for f in unmatched_flights[:15]:
            to_str = f["to_utc"].strftime("%H:%M UTC") if f["to_utc"] else "?"
            print(f"  {f['date']}  {f['dep']:>15} → {f['des']:<15}  "
                  f"T/O {to_str}  ({f['reg']})")


# Threshold sweep

def threshold_sweep(flights, score_thresholds):
    """Run matching at multiple score thresholds."""
    offshore = [f for f in flights if f["is_offshore"]]
    n_offshore = len(offshore)

    print(f"\n{'SCORE THRESHOLD SWEEP':-^70}")
    print(f"  {'Threshold':>9}  {'Events':>6}  {'Matched':>7}  {'Recall':>6}  "
          f"{'Events/Flight':>13}")

    for threshold in score_thresholds:
        events = load_detected_events(threshold)
        matched, unmatched_f, unmatched_e = match_flights_to_events(offshore, events)
        n_matched = len(matched)
        recall = n_matched / n_offshore * 100 if n_offshore > 0 else 0
        ratio = len(events) / n_matched if n_matched > 0 else float("inf")
        print(f"  {threshold:>9.0f}  {len(events):>6}  {n_matched:>7}  "
              f"{recall:>5.1f}%  {ratio:>13.1f}")


# Main

def print_canonical_report(flights, events, matched_pairs, matched_pair_keys,
                           matched_event_keys, unmatched_pairs, unmatched_events,
                           score_min, scope_label):
    """Report for canonical (strict + dwell + zero-tolerance) mode."""
    offshore_pairs = {(f["date"], f["reg"]) for f in flights if f.get("is_offshore")}
    n_off = len(offshore_pairs); n_match = len(matched_pair_keys)
    n_evt_match = len(matched_event_keys)
    fleet_events = [e for e in events if e["icao24"] in KNOWN_FLEET]
    n_evt_unmatch = len(unmatched_events)
    recall = n_match / n_off * 100 if n_off else 0
    prec = n_evt_match / (n_evt_match + n_evt_unmatch) * 100 if (n_evt_match + n_evt_unmatch) else 0
    f1 = 2 * prec * recall / (prec + recall) if (prec + recall) else 0
    print(f"\n{'═'*70}")
    print(f"  CANONICAL MODE — {scope_label}")
    print(f"  strict aircraft identity + inter-leg dwell + zero-tolerance overlap")
    print(f"  is_offshore v2 (DEP=DES > {SAME_AIRPORT_OFFSHORE_MIN_MINUTES} min counted as offshore round-trip)")
    print(f"  score ≥ {score_min:.0f}")
    print(f"{'═'*70}")
    print(f"  Offshore (date, tail) pairs: {n_off}")
    print(f"  Matched pairs:               {n_match}  ({recall:.1f}% recall)")
    print(f"  Unmatched pairs:             {len(unmatched_pairs)}")
    print(f"  Total emitted events ≥{score_min:.0f}: {len(events)}  (KNOWN_FLEET: {len(fleet_events)})")
    print(f"  Matched events:              {n_evt_match}  ({prec:.1f}% precision)")
    print(f"  Falsely detected (KNOWN_FLEET, no window match): {n_evt_unmatch}")
    print(f"  F1: {f1:.1f}%")


def main():
    parser = argparse.ArgumentParser(description="Validate helicopter detection")
    parser.add_argument("--score-min", type=float, default=40,
                        help="Minimum score threshold (default: 40 = τ_score event-emission cutoff)")
    parser.add_argument("--mode", choices=("canonical", "legacy"), default="canonical",
                        help="canonical: strict + inter-leg dwell + zero-tolerance overlap (default). "
                             "legacy: per-leg time-overlap with ±tolerance + KNOWN_FLEET filter only.")
    parser.add_argument("--scope", choices=("2024", "2025h1", "combined"), default="combined",
                        help="MFR + DB date scope (default: combined = 2024 + 2025 H1)")
    parser.add_argument("--sweep", action="store_true",
                        help="Run threshold sweep analysis (legacy mode only)")
    parser.add_argument("--by-sortie", action="store_true",
                        help="Match by sortie envelope (legacy mode only)")
    parser.add_argument("--sortie-tolerance", type=int, default=30,
                        help="Tolerance for sortie envelope matching (default: 30 min)")
    args = parser.parse_args()

    # Resolve scope → (loader, date_min, date_max, label)
    scope_map = {
        "2024":     (load_flights_2024,    dt.date(2024,1,1), dt.date(2025,1,1), "2024"),
        "2025h1":   (load_flights_2025_h1, dt.date(2025,1,1), dt.date(2025,7,1), "2025 H1"),
        "combined": (load_flights_combined, dt.date(2024,1,1), dt.date(2025,7,1), "2024 + 2025 H1"),
    }
    loader, dmin, dmax, scope_label = scope_map[args.scope]

    print("Loading flight report...")
    flights = loader()
    print(f"  {len(flights)} flights loaded ({scope_label})")

    print("Loading detected events...")
    events = load_detected_events(args.score_min, dmin, dmax)
    print(f"  {len(events)} events loaded at score ≥ {args.score_min:.0f}")

    if args.mode == "canonical":
        # Per-scope canonical report (split if combined for clarity)
        if args.scope == "combined":
            for sub_scope in ("2024", "2025h1"):
                sub_loader, sd_min, sd_max, sub_label = scope_map[sub_scope]
                sub_flights = sub_loader()
                sub_events = load_detected_events(args.score_min, sd_min, sd_max)
                mp, mpk, mek, ump, ume = match_flights_strict_dwell(sub_flights, sub_events)
                print_canonical_report(sub_flights, sub_events, mp, mpk, mek, ump, ume,
                                       args.score_min, sub_label)
        else:
            mp, mpk, mek, ump, ume = match_flights_strict_dwell(flights, events)
            print_canonical_report(flights, events, mp, mpk, mek, ump, ume,
                                   args.score_min, scope_label)
        return

    # Legacy mode (preserved for backwards comparison)
    print("Matching flights to events (legacy per-leg)...")
    matched, unmatched_flights, unmatched_events = match_flights_to_events(
        [f for f in flights if f["is_offshore"]], events
    )
    print_report(flights, events, matched, unmatched_flights, unmatched_events,
                 args.score_min)

    if args.by_sortie:
        print("\nGrouping legs into sorties...")
        sorties = group_legs_into_sorties(flights)
        print(f"  {len(sorties)} sorties built from {len(flights)} legs")
        s_matched, s_unmatched, s_unmatched_events = match_sorties_to_events(
            sorties, events, time_tolerance_min=args.sortie_tolerance)
        print_sortie_report(sorties, events, s_matched, s_unmatched,
                            s_unmatched_events, args.score_min)

    if args.sweep:
        threshold_sweep(flights, [10, 15, 20, 25, 30, 40, 50, 60, 70])


if __name__ == "__main__":
    main()
