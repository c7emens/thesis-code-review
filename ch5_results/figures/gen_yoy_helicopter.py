#!/usr/bin/env python3
"""Year-over-year helicopter activity: MFR destinations + detected events.

Generates two separate figures comparing 2024 to 2025 H1 on a per-month-rate
basis (no annualisation extrapolation, since H2 2025 is unknown):

  figures/data_stats/yoy_mfr_destinations.{pdf,png}
      MFR offshore-bound flights / month, bucketed by destination category.

  figures/data_stats/yoy_detected_events.{pdf,png}
      Stage 3 helicopter events / month at score >= 40 from known-fleet
      aircraft, bucketed by wind farm.

Source:
  - MFR sheets MVY/OQU 2024 (full year) and MVY/OQU 2025 (H1, dof < 2025-07-01)
  - stage3_helicopter_events DB table

Window discipline:
  - 2024 rate = total / 12.0
  - 2025 H1 rate = total / 6.0
Both rates are honest summaries of the data; no x2 projection.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import psycopg2

# Config

FLIGHT_REPORT_PRIMARY  = "/mnt/e/data_lake/helicopters/validation/Master Flight Report.xlsx"
FLIGHT_REPORT_FALLBACK = "/mnt/d/thesis/Master Flight Report.xlsx"
SHEETS_2024            = ["MVY 2024", "OQU 2024"]
SHEETS_2025_H1         = ["MVY 2025", "OQU 2025"]    # filtered to dof < 2025-07-01

DB = dict(host="localhost", port=5432, dbname="windfarm",
          user="thesis", password="thesis2026")

KNOWN_FLEET = {"a92f2d", "a932e4", "a9369b", "a93a52"}   # HSU helicopters
AIRPORT_CODES = {"KOQU", "KMVY", "KACK", "KACY", "KEWB", "KGON",
                 "KHYA", "KLDG", "KPVD", "MVY", "koqu", "6N5", "BKL1"}
SAME_AIRPORT_OFFSHORE_MIN_MINUTES = 30

FIG_DIR = Path("/mnt/d/thesis/main/figures/data_stats")
FIG_DIR.mkdir(parents=True, exist_ok=True)

# Window months for per-month rate normalisation
MONTHS_2024    = 12.0
MONTHS_2025_H1 = 6.0


def _resolve_mfr() -> str:
    try:
        if Path(FLIGHT_REPORT_PRIMARY).exists():
            return FLIGHT_REPORT_PRIMARY
    except OSError:
        pass
    return FLIGHT_REPORT_FALLBACK


# MFR loading

HEADER_LABELS = {
    "reg": "A/C REG", "mission": "Mission",
    "dep": "DEP", "des": "DES",
    "dof": "DOF", "to": "T/O", "ldg": "LDG",
}


def _header_map(ws) -> dict[str, int]:
    header = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    return {k: i for k, lbl in HEADER_LABELS.items()
              for i, c in enumerate(header) if c == lbl}


def _parse_time(val) -> dt.time | None:
    if isinstance(val, dt.time):
        return val
    if isinstance(val, dt.timedelta):
        t = int(val.total_seconds())
        return dt.time((t // 3600) % 24, (t // 60) % 60)
    if isinstance(val, str):
        val = val.strip()
        if ":" in val:
            try:
                h, m = val.split(":")[:2]
                return dt.time(int(h), int(m))
            except (ValueError, IndexError):
                return None
    return None


def load_mfr_flights(sheets, dof_max=None) -> list[dict]:
    """Return all offshore-bound MFR rows from the given sheets."""
    wb = openpyxl.load_workbook(_resolve_mfr(), read_only=True, data_only=True)
    out = []
    for sheet in sheets:
        ws = wb[sheet]
        cols = _header_map(ws)
        if "dof" not in cols or "dep" not in cols or "des" not in cols:
            continue
        for row in ws.iter_rows(min_row=3, values_only=True):
            def g(k):
                i = cols.get(k)
                return row[i] if i is not None and i < len(row) else None
            dof, dep, des = g("dof"), g("dep"), g("des")
            if dof is None or dep is None or des is None:
                continue
            if not isinstance(dof, dt.datetime):
                continue
            if dof_max is not None and dof.date() >= dof_max:
                continue
            dep, des = str(dep).strip(), str(des).strip()

            # Offshore classification (v2 rule)
            offshore = (dep not in AIRPORT_CODES) or (des not in AIRPORT_CODES)
            if not offshore and dep == des:
                to_t, ldg_t = _parse_time(g("to")), _parse_time(g("ldg"))
                if to_t and ldg_t:
                    dur = (dt.datetime.combine(dof.date(), ldg_t)
                           - dt.datetime.combine(dof.date(), to_t)).total_seconds() / 60
                    offshore = dur >= SAME_AIRPORT_OFFSHORE_MIN_MINUTES
            if not offshore:
                continue

            out.append({
                "dof": dof.date(),
                "dep": dep,
                "des": des,
                "mission": str(g("mission")).strip() if g("mission") else "",
            })
    return out


# Destination bucketing

def bucket_destination(dep: str, des: str, mission: str) -> str:
    """Return the destination category label for one MFR flight row.

    Decide which named MFR destinations belong in each bucket based on
    your domain knowledge of the HSU fleet's offshore operations during
    2024–2025 H1. Buckets feed the x-axis of the MFR destinations figure.

    Suggested categories (you can change them):
      - "Construction vessels" — e.g. SeaJacks Scylla, Bokalift, Orion,
        ESP, generic "Installation Vessels"
      - "SOV / O&M platforms"  — e.g. Leviathan, Eco Edison, HEA Leviathan
      - "Other offshore"        — named destinations that don't fit above
      - "Same-airport round-trip" — DEP == DES with offshore-classified duration
    """
    # Same-airport offshore round-trip (DEP == DES, already validated as
    # offshore-duration in load_mfr_flights via the v2 rule).
    if dep == des:
        return "Same-airport round-trip"

    d = (des or "").lower()

    # Construction / installation vessels: jack-ups, heavy-lift cranes,
    # generic "installation vessels" umbrella entries.
    if any(kw in d for kw in (
        "scylla", "bokalift", "orion", "installation",
        "sea installer", "aeolus", "sea challenger",
    )):
        return "Construction vessels"

    # SOV / O&M platforms: walk-to-work service-operations vessels.
    if any(kw in d for kw in ("leviathan", "eco edison", "ecoedison")):
        return "SOV / O&M platforms"

    return "Other offshore"


# DB loading

def load_events_by_farm(year_filter: tuple[dt.date, dt.date]) -> dict[str, int]:
    """Return {project_name: count} of score>=40 events from KNOWN_FLEET
    in the given date range [start, end)."""
    start, end = year_filter
    fleet = tuple(KNOWN_FLEET)
    with psycopg2.connect(**DB) as conn, conn.cursor() as cur:
        cur.execute("""
            SELECT project_name, COUNT(*)
            FROM stage3_helicopter_events
            WHERE score >= 40
              AND icao24 IN %s
              AND visit_start >= %s
              AND visit_start <  %s
            GROUP BY project_name
            ORDER BY project_name
        """, (fleet, start, end))
        return dict(cur.fetchall())


# Plots

BAR_2024_COLOR = "#1f77b4"
BAR_2025_COLOR = "#ff7f0e"


def plot_mfr_destinations(buckets_2024: dict[str, int],
                          buckets_2025: dict[str, int],
                          out_stem: str):
    """Grouped-bar plot of MFR flights / month per destination bucket."""
    cats = sorted(set(buckets_2024) | set(buckets_2025))
    rates_2024 = [buckets_2024.get(c, 0) / MONTHS_2024 for c in cats]
    rates_2025 = [buckets_2025.get(c, 0) / MONTHS_2025_H1 for c in cats]

    x = np.arange(len(cats))
    width = 0.38
    fig, ax = plt.subplots(figsize=(10, 5.5))
    ax.bar(x - width/2, rates_2024, width, color=BAR_2024_COLOR,
           label=f"2024 (n / {int(MONTHS_2024)} mo)", edgecolor="black", linewidth=0.4)
    ax.bar(x + width/2, rates_2025, width, color=BAR_2025_COLOR,
           label=f"2025 H1 (n / {int(MONTHS_2025_H1)} mo)", edgecolor="black", linewidth=0.4)
    ax.set_xticks(x)
    ax.set_xticklabels(cats, rotation=15, ha="right", fontsize=10)
    ax.set_ylabel("MFR offshore flights per month", fontsize=11)
    ax.set_title("MFR helicopter offshore destinations — per-month rate", fontsize=12)
    ax.tick_params(axis="y", labelsize=10)
    # Legend in upper-right INSIDE plot but lifted above bars; ymax buffer avoids overlap.
    ymax = max(max(rates_2024, default=0), max(rates_2025, default=0))
    ax.set_ylim(0, ymax * 1.25)
    ax.legend(loc="upper right", framealpha=0.95, fontsize=10)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    plt.tight_layout()
    for ext in ("pdf", "png"):
        fig.savefig(FIG_DIR / f"{out_stem}.{ext}", dpi=160, bbox_inches="tight")
    plt.close(fig)
    print(f"wrote {FIG_DIR / (out_stem + '.pdf')}")


def plot_detected_events(events_2024: dict[str, int],
                         events_2025: dict[str, int],
                         out_stem: str):
    """Grouped-bar plot of detected events / month per wind farm."""
    farms = sorted(set(events_2024) | set(events_2025))
    rates_2024 = [events_2024.get(f, 0) / MONTHS_2024 for f in farms]
    rates_2025 = [events_2025.get(f, 0) / MONTHS_2025_H1 for f in farms]

    x = np.arange(len(farms))
    width = 0.38
    fig, ax = plt.subplots(figsize=(10, 5.5))
    ax.bar(x - width/2, rates_2024, width, color=BAR_2024_COLOR,
           label=f"2024 (n / {int(MONTHS_2024)} mo)", edgecolor="black", linewidth=0.4)
    ax.bar(x + width/2, rates_2025, width, color=BAR_2025_COLOR,
           label=f"2025 H1 (n / {int(MONTHS_2025_H1)} mo)", edgecolor="black", linewidth=0.4)
    ax.set_xticks(x)
    ax.set_xticklabels([f.replace("_", " ") for f in farms],
                       rotation=15, ha="right", fontsize=10)
    ax.set_ylabel("Detected events per month  (score $\\geq 40$, known fleet)", fontsize=11)
    ax.set_title("Detected helicopter events — per-month rate by farm", fontsize=12)
    ax.tick_params(axis="y", labelsize=10)
    ymax = max(max(rates_2024, default=0), max(rates_2025, default=0))
    ax.set_ylim(0, ymax * 1.25)
    ax.legend(loc="upper right", framealpha=0.95, fontsize=10)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    plt.tight_layout()
    for ext in ("pdf", "png"):
        fig.savefig(FIG_DIR / f"{out_stem}.{ext}", dpi=160, bbox_inches="tight")
    plt.close(fig)
    print(f"wrote {FIG_DIR / (out_stem + '.pdf')}")


# Main

def main():
    print("Loading MFR flights...")
    flights_2024 = load_mfr_flights(SHEETS_2024)
    flights_2025 = load_mfr_flights(SHEETS_2025_H1, dof_max=dt.date(2025, 7, 1))
    print(f"  2024: {len(flights_2024)} offshore flights")
    print(f"  2025 H1: {len(flights_2025)} offshore flights")

    buckets_2024: dict[str, int] = {}
    buckets_2025: dict[str, int] = {}
    for f in flights_2024:
        b = bucket_destination(f["dep"], f["des"], f["mission"])
        buckets_2024[b] = buckets_2024.get(b, 0) + 1
    for f in flights_2025:
        b = bucket_destination(f["dep"], f["des"], f["mission"])
        buckets_2025[b] = buckets_2025.get(b, 0) + 1

    print("Destination buckets:")
    for cat in sorted(set(buckets_2024) | set(buckets_2025)):
        print(f"  {cat:30s}  2024={buckets_2024.get(cat, 0):>4}   2025 H1={buckets_2025.get(cat, 0):>4}")

    print("Loading detected events from DB...")
    events_2024 = load_events_by_farm((dt.date(2024, 1, 1), dt.date(2025, 1, 1)))
    events_2025 = load_events_by_farm((dt.date(2025, 1, 1), dt.date(2025, 7, 1)))
    print(f"  2024 events: {sum(events_2024.values())}")
    print(f"  2025 H1 events: {sum(events_2025.values())}")

    plot_mfr_destinations(buckets_2024, buckets_2025, "yoy_mfr_destinations")
    plot_detected_events(events_2024, events_2025, "yoy_detected_events")


if __name__ == "__main__":
    main()
